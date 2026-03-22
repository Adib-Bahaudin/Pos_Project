from datetime import datetime
import importlib.util
from pathlib import Path

from PySide6.QtCore import QDate, QTimer, Qt
from PySide6.QtGui import QColor, QFont, QTextDocument
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDateEdit,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from database import DatabaseManager


class SejarahTransaksiWindow(QWidget):
    DEFAULT_ITEMS_PER_PAGE = 20
    SORT_FIELDS = {
        0: "id",
        1: "tanggal",
        2: "kasir",
        3: "customer",
        4: "items",
        5: "total",
        6: "metode",
    }
    PAYMENT_COLORS = {
        "Tunai": "#123524",
        "QRIS": "#093b4a",
        "Transfer": "#102c68",
        "Kartu": "#3f2b68",
        "E-Wallet": "#4b3a0a",
    }

    def __init__(self, user_data=None, db_manager=None):
        super().__init__()
        self.user_data = user_data or {}
        self.db_manager = db_manager or DatabaseManager("db_BarokahCopy.db")
        self.current_page = 1
        self.total_records = 0
        self.current_transactions = []
        self.current_statistics = {}
        self.current_filters = {}
        self.current_sort_by = "tanggal"
        self.current_sort_direction = "DESC"
        self.is_loading = False

        self.setup_ui()
        self.load_filter_options()
        QTimer.singleShot(0, self.apply_filters)

    def setup_ui(self):
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        container = QFrame()
        container.setStyleSheet("background-color: #050505;")
        root_layout.addWidget(container)

        content_layout = QVBoxLayout(container)
        content_layout.setContentsMargins(24, 24, 24, 24)
        content_layout.setSpacing(18)

        content_layout.addWidget(self._create_header_card())
        content_layout.addWidget(self._create_statistics_panel())
        content_layout.addWidget(self._create_filter_panel())
        content_layout.addWidget(self._create_table_panel(), 1)
        content_layout.addLayout(self._create_pagination_panel())
        content_layout.addLayout(self._create_action_panel())

        self.setStyleSheet(self._get_stylesheet())

    def _create_header_card(self):
        card = self._build_card()
        layout = QHBoxLayout(card)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(16)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(4)

        title = QLabel("SEJARAH TRANSAKSI")
        title.setFont(QFont("Times New Roman", 22, QFont.Weight.Bold))
        title.setStyleSheet("color: #ffffff;")

        subtitle = QLabel("Pantau histori penjualan, filter data, dan ekspor laporan transaksi kasir.")
        subtitle.setStyleSheet("color: #98a3af; font-size: 12px;")

        user_name = str(self.user_data.get("username") or "Guest").upper()
        role_name = str(self.user_data.get("role") or "user").replace("_", " ").title()
        badge = QLabel(f"Kasir Aktif: {user_name} • Role: {role_name}")
        badge.setObjectName("smallBadge")

        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)
        title_layout.addWidget(badge, alignment=Qt.AlignmentFlag.AlignLeft)

        layout.addLayout(title_layout)
        layout.addStretch()
        return card

    def _create_statistics_panel(self):
        wrapper = QWidget()
        layout = QHBoxLayout(wrapper)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(14)

        self.metric_total_trx = self._create_metric_card("Total Transaksi", "0", "Jumlah transaksi sesuai filter")
        self.metric_revenue = self._create_metric_card("Total Revenue", "Rp 0", "Akumulasi omset")
        self.metric_avg = self._create_metric_card("Rata-rata Transaksi", "Rp 0", "Nilai rata-rata per transaksi")
        self.metric_top_cashier = self._create_metric_card("Kasir Teratas", "-", "Kasir paling aktif")

        for card in [self.metric_total_trx, self.metric_revenue, self.metric_avg, self.metric_top_cashier]:
            layout.addWidget(card)

        return wrapper

    def _create_metric_card(self, title_text, value_text, description_text):
        card = self._build_card()
        card.setMinimumHeight(120)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(6)

        title = QLabel(title_text)
        title.setStyleSheet("color: #98a3af; font-size: 12px; font-weight: 600;")

        value = QLabel(value_text)
        value.setStyleSheet("color: #ffffff; font-size: 22px; font-weight: 800;")

        description = QLabel(description_text)
        description.setStyleSheet("color: #6b7a8c; font-size: 11px;")
        description.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(value)
        layout.addStretch()
        layout.addWidget(description)

        card.metric_value = value
        card.metric_description = description
        return card

    def _create_filter_panel(self):
        card = self._build_card()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        title = QLabel("Filter Histori")
        title.setStyleSheet("color: #ffffff; font-size: 15px; font-weight: 700;")
        layout.addWidget(title)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)

        self.date_from_input = QDateEdit()
        self.date_from_input.setCalendarPopup(True)
        self.date_from_input.setDate(QDate.currentDate().addMonths(-1))
        self.date_from_input.setDisplayFormat("dd MMM yyyy")

        self.date_to_input = QDateEdit()
        self.date_to_input.setCalendarPopup(True)
        self.date_to_input.setDate(QDate.currentDate())
        self.date_to_input.setDisplayFormat("dd MMM yyyy")

        self.kasir_filter = QComboBox()
        self.kasir_filter.addItem("Semua Kasir", None)

        self.payment_filter = QComboBox()
        self.payment_filter.addItem("Semua Metode")
        self.payment_filter.addItems(["Tunai", "QRIS", "Transfer", "Kartu", "E-Wallet"])

        self.amount_min_input = QLineEdit()
        self.amount_min_input.setPlaceholderText("Min. total")

        self.amount_max_input = QLineEdit()
        self.amount_max_input.setPlaceholderText("Maks. total")

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari nama customer / ID transaksi / item...")
        self.search_input.returnPressed.connect(self.apply_filters)

        self.button_filter = QPushButton("FILTER")
        self.button_filter.setObjectName("primaryButton")
        self.button_filter.clicked.connect(self.apply_filters)

        grid.addWidget(self._create_form_label("Dari"), 0, 0)
        grid.addWidget(self.date_from_input, 0, 1)
        grid.addWidget(self._create_form_label("Sampai"), 0, 2)
        grid.addWidget(self.date_to_input, 0, 3)
        grid.addWidget(self._create_form_label("Kasir"), 1, 0)
        grid.addWidget(self.kasir_filter, 1, 1)
        grid.addWidget(self._create_form_label("Metode Bayar"), 1, 2)
        grid.addWidget(self.payment_filter, 1, 3)
        grid.addWidget(self._create_form_label("Amount Range"), 2, 0)
        grid.addWidget(self.amount_min_input, 2, 1)
        grid.addWidget(self.amount_max_input, 2, 2)
        grid.addWidget(self.search_input, 2, 3)

        action_row = QHBoxLayout()
        action_row.addStretch()
        action_row.addWidget(self.button_filter)

        self.filter_info_label = QLabel("Gunakan filter untuk memuat histori transaksi.")
        self.filter_info_label.setStyleSheet("color: #98a3af; font-size: 12px;")

        layout.addLayout(grid)
        layout.addWidget(self.filter_info_label)
        layout.addLayout(action_row)
        return card

    def _create_table_panel(self):
        card = self._build_card()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        title = QLabel("Daftar Transaksi")
        title.setStyleSheet("color: #ffffff; font-size: 15px; font-weight: 700;")
        layout.addWidget(title)

        self.transaction_table = QTableWidget(0, 8)
        self.transaction_table.setHorizontalHeaderLabels([
            "ID", "Tanggal", "Kasir", "Customer", "Items", "Total", "Metode", "Action"
        ])
        self.transaction_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.transaction_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.transaction_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.transaction_table.setAlternatingRowColors(False)
        self.transaction_table.setSortingEnabled(False)
        self.transaction_table.setWordWrap(True)
        self.transaction_table.verticalHeader().setVisible(False)
        self.transaction_table.verticalHeader().setDefaultSectionSize(54)
        self.transaction_table.horizontalHeader().setStretchLastSection(False)
        self.transaction_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.transaction_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.transaction_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.transaction_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.transaction_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.transaction_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.transaction_table.horizontalHeader().sectionClicked.connect(self._handle_sort_change)
        self.transaction_table.cellDoubleClicked.connect(self._handle_row_open)

        self.empty_state_label = QLabel("Tidak ada transaksi")
        self.empty_state_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.empty_state_label.setStyleSheet("color: #98a3af; font-size: 13px; padding: 24px;")
        self.empty_state_label.hide()

        layout.addWidget(self.transaction_table)
        layout.addWidget(self.empty_state_label)
        return card

    def _create_pagination_panel(self):
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        self.button_prev = QPushButton("← Previous")
        self.button_prev.setObjectName("ghostButton")
        self.button_prev.clicked.connect(lambda: self.on_pagination_change(self.current_page - 1))

        self.page_label = QLabel("Halaman 1 / 1")
        self.page_label.setStyleSheet("color: #ffffff; font-size: 12px; font-weight: 600;")

        self.button_next = QPushButton("Next →")
        self.button_next.setObjectName("ghostButton")
        self.button_next.clicked.connect(lambda: self.on_pagination_change(self.current_page + 1))

        self.items_per_page_combo = QComboBox()
        self.items_per_page_combo.addItems(["10", "20", "50", "100"])
        self.items_per_page_combo.setCurrentText(str(self.DEFAULT_ITEMS_PER_PAGE))
        self.items_per_page_combo.currentTextChanged.connect(self._on_items_per_page_changed)

        layout.addWidget(self.button_prev)
        layout.addWidget(self.page_label)
        layout.addWidget(self.button_next)
        layout.addStretch()
        layout.addWidget(self._create_form_label("Items/Page"))
        layout.addWidget(self.items_per_page_combo)
        return layout

    def _create_action_panel(self):
        layout = QHBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        self.button_print = QPushButton("🖨️ Print")
        self.button_print.setObjectName("ghostButton")
        self.button_print.clicked.connect(self._handle_print_action)

        self.button_export_excel = QPushButton("📥 Export Excel")
        self.button_export_excel.setObjectName("ghostButton")
        self.button_export_excel.clicked.connect(lambda: self.export_to_excel(self.current_transactions))

        self.button_export_pdf = QPushButton("📄 Export PDF")
        self.button_export_pdf.setObjectName("ghostButton")
        self.button_export_pdf.clicked.connect(lambda: self.export_to_pdf(self.current_transactions))

        self.button_refresh = QPushButton("🔄 Refresh")
        self.button_refresh.setObjectName("primaryButton")
        self.button_refresh.clicked.connect(self.apply_filters)

        layout.addWidget(self.button_print)
        layout.addWidget(self.button_export_excel)
        layout.addWidget(self.button_export_pdf)
        layout.addStretch()
        layout.addWidget(self.button_refresh)
        return layout

    def _build_card(self):
        card = QFrame()
        card.setObjectName("historyCard")
        card.setFrameShape(QFrame.Shape.NoFrame)
        return card

    def _create_form_label(self, text):
        label = QLabel(text)
        label.setObjectName("formLabel")
        return label

    def _get_stylesheet(self):
        return """
            QFrame#historyCard {
                background-color: #101820;
                border: 1px solid #1f2b38;
                border-radius: 16px;
            }
            QLabel#formLabel {
                color: #b8c4d0;
                font-size: 12px;
                font-weight: 600;
            }
            QLabel#smallBadge {
                color: #dbe8f4;
                background-color: #17202a;
                border: 1px solid #2a3745;
                border-radius: 10px;
                padding: 6px 12px;
                font-size: 12px;
                font-weight: 600;
            }
            QLineEdit, QComboBox, QDateEdit {
                background-color: #091119;
                color: #ffffff;
                border: 1px solid #233243;
                border-radius: 10px;
                padding: 8px 12px;
                min-height: 20px;
                font-size: 12px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border: 1px solid #1c8adb;
            }
            QTableWidget {
                background-color: #0a1219;
                color: #ffffff;
                border: 1px solid #1c2833;
                border-radius: 12px;
                gridline-color: #16212a;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #112030;
                color: #dbe8f4;
                border: none;
                padding: 10px;
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton {
                border-radius: 10px;
                padding: 10px 14px;
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton#primaryButton {
                background-color: #157347;
                color: #ffffff;
                border: 1px solid #1e9f63;
            }
            QPushButton#primaryButton:hover {
                background-color: #1e9f63;
            }
            QPushButton#ghostButton {
                background-color: #17202a;
                color: #dbe8f4;
                border: 1px solid #2a3745;
            }
            QPushButton#ghostButton:hover {
                background-color: #233243;
            }
        """

    def load_filter_options(self):
        self.kasir_filter.blockSignals(True)
        while self.kasir_filter.count() > 1:
            self.kasir_filter.removeItem(1)

        current_user_id = self.user_data.get("user_id") or self.user_data.get("id")
        current_role = str(self.user_data.get("role") or "")
        for cashier in self.db_manager.get_cashier_list():
            self.kasir_filter.addItem(cashier.get("nama", "-"), cashier.get("id"))
            if cashier.get("id") == current_user_id and current_role != "Super_user":
                self.kasir_filter.setCurrentIndex(self.kasir_filter.count() - 1)

        if current_role != "Super_user":
            self.kasir_filter.setEnabled(False)
        self.kasir_filter.blockSignals(False)

    def _validate_filter_inputs(self):
        if self.date_from_input.date() > self.date_to_input.date():
            raise ValueError("Tanggal 'dari' tidak boleh lebih besar dari 'sampai'.")

        amount_min = self._parse_amount(self.amount_min_input.text().strip())
        amount_max = self._parse_amount(self.amount_max_input.text().strip())
        if amount_min is not None and amount_max is not None and amount_min > amount_max:
            raise ValueError("Amount minimum tidak boleh lebih besar dari amount maksimum.")
        return amount_min, amount_max

    @staticmethod
    def _parse_amount(value):
        cleaned = "".join(char for char in str(value) if char.isdigit())
        return int(cleaned) if cleaned else None

    def _build_filters(self):
        amount_min, amount_max = self._validate_filter_inputs()
        filters = {
            "date_from": self.date_from_input.date().toString("yyyy-MM-dd"),
            "date_to": self.date_to_input.date().toString("yyyy-MM-dd"),
            "kasir_id": self.kasir_filter.currentData(),
            "payment_method": self.payment_filter.currentText(),
            "amount_min": amount_min,
            "amount_max": amount_max,
            "search_keyword": self.search_input.text().strip(),
            "limit": int(self.items_per_page_combo.currentText()),
            "offset": (self.current_page - 1) * int(self.items_per_page_combo.currentText()),
            "sort_by": self.current_sort_by,
            "sort_direction": self.current_sort_direction,
        }
        if str(self.user_data.get("role") or "") != "Super_user":
            filters["user_scope_id"] = self.user_data.get("user_id") or self.user_data.get("id")
        return filters

    def set_loading_state(self, is_loading, message=None):
        self.is_loading = is_loading
        for widget in [
            self.button_filter,
            self.button_refresh,
            self.button_export_excel,
            self.button_export_pdf,
            self.button_print,
        ]:
            widget.setDisabled(is_loading)

        if message:
            self.filter_info_label.setText(message)

    def fetch_transactions(self, filters_dict):
        result = self.db_manager.get_transaction_history(filters_dict)
        self.total_records = result.get("total_count", 0)
        return result.get("rows", [])

    def fetch_transaction_detail(self, transaction_id):
        return self.db_manager.get_transaction_detail_with_items(transaction_id)

    def get_statistics(self, filters_dict):
        return self.db_manager.get_transaction_statistics(filters_dict)

    def populate_table(self, transactions_data):
        self.transaction_table.setRowCount(0)
        self.empty_state_label.setVisible(not transactions_data)
        self.transaction_table.setVisible(bool(transactions_data))

        for row_index, transaction in enumerate(transactions_data):
            self.transaction_table.insertRow(row_index)

            values = [
                str(transaction.get("id", "-")),
                self._format_datetime(transaction.get("tanggal")),
                transaction.get("kasir_name", "-"),
                transaction.get("customer_name", "Pelanggan Umum"),
                transaction.get("items_preview", "-"),
                self._format_currency(transaction.get("total", 0)),
                transaction.get("payment_method", "-"),
            ]

            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setData(Qt.ItemDataRole.UserRole, transaction.get("id"))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter if column in [0, 5, 6] else Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                self.transaction_table.setItem(row_index, column, item)

            button = QPushButton("Detail")
            button.setObjectName("ghostButton")
            button.clicked.connect(lambda _, trx_id=transaction.get("id"): self.show_transaction_detail_modal(trx_id))
            self.transaction_table.setCellWidget(row_index, 7, button)

            self._apply_row_payment_color(row_index, transaction.get("payment_method"))

        self.transaction_table.resizeRowsToContents()

    def _apply_row_payment_color(self, row_index, payment_method):
        color = QColor(self.PAYMENT_COLORS.get(str(payment_method), "#1a1f24"))
        for column in range(7):
            item = self.transaction_table.item(row_index, column)
            if item:
                item.setBackground(color)

    def show_transaction_detail_modal(self, transaction_id):
        detail = self.fetch_transaction_detail(transaction_id)
        if not detail:
            QMessageBox.warning(self, "Detail Transaksi", "Detail transaksi tidak ditemukan.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Detail Transaksi #{transaction_id}")
        dialog.setModal(True)
        dialog.resize(900, 620)
        dialog.setStyleSheet(self._get_stylesheet())

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        header = detail.get("header", {})
        header_text = QTextEdit()
        header_text.setReadOnly(True)
        header_text.setMinimumHeight(190)
        header_text.setHtml(self._build_detail_html(header))

        items_table = QTableWidget(0, 5)
        items_table.setHorizontalHeaderLabels(["SKU", "Produk", "Qty", "Harga", "Subtotal"])
        items_table.verticalHeader().setVisible(False)
        items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        items_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        items_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        items_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        items_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        for row_index, item in enumerate(detail.get("items", [])):
            items_table.insertRow(row_index)
            row_values = [
                item.get("sku", "-"),
                item.get("product_name", "-"),
                str(item.get("quantity", 0)),
                self._format_currency(item.get("price", 0)),
                self._format_currency(item.get("subtotal", 0)),
            ]
            for column, value in enumerate(row_values):
                table_item = QTableWidgetItem(value)
                if column in [2, 3, 4]:
                    table_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                items_table.setItem(row_index, column, table_item)

        button_row = QHBoxLayout()
        button_row.addStretch()
        print_button = QPushButton("Print")
        print_button.setObjectName("primaryButton")
        print_button.clicked.connect(lambda: self.print_transaction(transaction_id))
        close_button = QPushButton("Close")
        close_button.setObjectName("ghostButton")
        close_button.clicked.connect(dialog.accept)
        button_row.addWidget(print_button)
        button_row.addWidget(close_button)

        layout.addWidget(header_text)
        layout.addWidget(items_table, 1)
        layout.addLayout(button_row)
        dialog.exec()

    def _build_detail_html(self, header):
        return f"""
            <div style='color:#dbe8f4; font-size:12px;'>
                <h2 style='color:#ffffff;'>Transaksi #{header.get('id', '-')}</h2>
                <p><b>Tanggal:</b> {self._format_datetime(header.get('tanggal'))}</p>
                <p><b>Kasir:</b> {header.get('kasir_name', '-')}</p>
                <p><b>Customer:</b> {header.get('customer_name', 'Pelanggan Umum')}</p>
                <p><b>Metode Bayar:</b> {header.get('payment_method', '-')}</p>
                <p><b>Subtotal:</b> {self._format_currency(header.get('subtotal', 0))} &nbsp; | &nbsp;
                   <b>Diskon:</b> {self._format_currency(header.get('diskon_nominal', 0))} ({header.get('diskon_persen', 0)}%) &nbsp; | &nbsp;
                   <b>Pembulatan:</b> {self._format_currency(header.get('pembulatan', 0))}</p>
                <p><b>Total:</b> {self._format_currency(header.get('total', 0))} &nbsp; | &nbsp;
                   <b>Dibayar:</b> {self._format_currency(header.get('amount_paid', 0))} &nbsp; | &nbsp;
                   <b>Kembalian:</b> {self._format_currency(header.get('change_amount', 0))}</p>
                <p><b>Laba Bersih:</b> {self._format_currency(header.get('laba_bersih', 0))} &nbsp; | &nbsp;
                   <b>Total HPP:</b> {self._format_currency(header.get('total_hpp', 0))}</p>
                <p><b>Catatan:</b> {header.get('notes', '-') or '-'}</p>
            </div>
        """

    def apply_filters(self):
        if self.is_loading:
            return

        try:
            self.current_page = 1
            filters = self._build_filters()
        except ValueError as error:
            QMessageBox.warning(self, "Validasi Filter", str(error))
            return

        self._load_transactions(filters, reset_page=True)

    def _load_transactions(self, filters, reset_page=False):
        try:
            self.set_loading_state(True, "Memuat histori transaksi...")
            if reset_page:
                filters["offset"] = 0
            self.current_filters = dict(filters)
            self.current_transactions = self.fetch_transactions(filters)
            self.current_statistics = self.get_statistics(filters)
            self.populate_table(self.current_transactions)
            self._update_statistics_cards(self.current_statistics)
            self._update_pagination_controls()

            loaded = len(self.current_transactions)
            self.filter_info_label.setText(
                f"Menampilkan {loaded} transaksi dari total {self.total_records} data."
            )
        except Exception as error:
            QMessageBox.critical(self, "Database Error", f"Gagal memuat histori transaksi:\n{error}")
            self.filter_info_label.setText("Terjadi kesalahan saat memuat data transaksi.")
        finally:
            self.set_loading_state(False)

    def _update_statistics_cards(self, stats):
        self.metric_total_trx.metric_value.setText(str(stats.get("total_count", 0)))
        self.metric_revenue.metric_value.setText(self._format_currency(stats.get("total_revenue", 0)))
        self.metric_avg.metric_value.setText(self._format_currency(stats.get("avg_transaction", 0)))
        top_cashier = stats.get("top_cashier", {})
        top_text = top_cashier.get("name", "-")
        if top_cashier.get("count"):
            top_text = f"{top_text} ({top_cashier.get('count', 0)} trx)"
        self.metric_top_cashier.metric_value.setText(top_text)

    def on_pagination_change(self, page_num):
        total_pages = max(1, (self.total_records + int(self.items_per_page_combo.currentText()) - 1) // int(self.items_per_page_combo.currentText()))
        new_page = max(1, min(page_num, total_pages))
        if new_page == self.current_page and self.current_transactions:
            return

        self.current_page = new_page
        filters = dict(self.current_filters or self._build_filters())
        filters["limit"] = int(self.items_per_page_combo.currentText())
        filters["offset"] = (self.current_page - 1) * filters["limit"]
        self._load_transactions(filters)

    def _update_pagination_controls(self):
        per_page = int(self.items_per_page_combo.currentText())
        total_pages = max(1, (self.total_records + per_page - 1) // per_page)
        self.page_label.setText(f"Halaman {self.current_page} / {total_pages}")
        self.button_prev.setEnabled(self.current_page > 1)
        self.button_next.setEnabled(self.current_page < total_pages)

    def _on_items_per_page_changed(self, *_args):
        self.current_page = 1
        self.apply_filters()

    def _handle_sort_change(self, section_index):
        if section_index not in self.SORT_FIELDS:
            return

        selected_field = self.SORT_FIELDS[section_index]
        if self.current_sort_by == selected_field:
            self.current_sort_direction = "ASC" if self.current_sort_direction == "DESC" else "DESC"
        else:
            self.current_sort_by = selected_field
            self.current_sort_direction = "ASC" if selected_field in {"id", "tanggal", "total"} else "DESC"
        self.apply_filters()

    def _handle_row_open(self, row, _column):
        item = self.transaction_table.item(row, 0)
        if not item:
            return
        transaction_id = item.data(Qt.ItemDataRole.UserRole)
        if transaction_id:
            self.show_transaction_detail_modal(transaction_id)

    def export_to_excel(self, transactions_data):
        if not transactions_data:
            QMessageBox.information(self, "Export Excel", "Tidak ada transaksi untuk diekspor.")
            return
        if importlib.util.find_spec("openpyxl") is None:
            QMessageBox.warning(
                self,
                "Export Excel",
                "Package openpyxl belum tersedia di environment ini.",
            )
            return

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        default_name = self._build_export_filename("xlsx")
        file_path, _ = QFileDialog.getSaveFileName(self, "Simpan Excel", default_name, "Excel Workbook (*.xlsx)")
        if not file_path:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sejarah Transaksi"

        sheet["A1"] = "Laporan Sejarah Transaksi"
        sheet["A1"].font = Font(bold=True, size=16)
        sheet["A2"] = f"Periode: {self.current_filters.get('date_from', '-')} s/d {self.current_filters.get('date_to', '-')}"
        sheet["A3"] = f"Kasir: {self.kasir_filter.currentText()}"
        sheet["A4"] = f"Metode Bayar: {self.payment_filter.currentText()}"

        stats = self.current_statistics or {}
        sheet["F1"] = "Statistik"
        sheet["F1"].font = Font(bold=True, size=14)
        sheet["F2"] = "Total Transaksi"
        sheet["G2"] = stats.get("total_count", 0)
        sheet["F3"] = "Total Revenue"
        sheet["G3"] = stats.get("total_revenue", 0)
        sheet["F4"] = "Rata-rata"
        sheet["G4"] = stats.get("avg_transaction", 0)
        top_cashier = stats.get("top_cashier", {})
        sheet["F5"] = "Kasir Teratas"
        sheet["G5"] = f"{top_cashier.get('name', '-')} ({top_cashier.get('count', 0)} trx)"

        headers = ["ID", "Tanggal", "Kasir", "Customer", "Items", "Total", "Metode"]
        header_row = 7
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True, color="FFFFFF")

        for row_offset, transaction in enumerate(transactions_data, start=1):
            sheet.cell(row=header_row + row_offset, column=1, value=transaction.get("id"))
            sheet.cell(row=header_row + row_offset, column=2, value=self._format_datetime(transaction.get("tanggal")))
            sheet.cell(row=header_row + row_offset, column=3, value=transaction.get("kasir_name"))
            sheet.cell(row=header_row + row_offset, column=4, value=transaction.get("customer_name"))
            sheet.cell(row=header_row + row_offset, column=5, value=transaction.get("items_preview"))
            total_cell = sheet.cell(row=header_row + row_offset, column=6, value=int(transaction.get("total", 0)))
            total_cell.number_format = '"Rp" #,##0'
            sheet.cell(row=header_row + row_offset, column=7, value=transaction.get("payment_method"))

        for col_idx, width in enumerate([12, 24, 18, 20, 40, 18, 16], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        workbook.save(file_path)
        QMessageBox.information(self, "Export Excel", f"File Excel berhasil disimpan ke:\n{file_path}")

    def export_to_pdf(self, transactions_data):
        if not transactions_data:
            QMessageBox.information(self, "Export PDF", "Tidak ada transaksi untuk diekspor.")
            return
        if importlib.util.find_spec("reportlab") is None:
            QMessageBox.warning(
                self,
                "Export PDF",
                "Package reportlab belum tersedia di environment ini.",
            )
            return

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        default_name = self._build_export_filename("pdf")
        file_path, _ = QFileDialog.getSaveFileName(self, "Simpan PDF", default_name, "PDF File (*.pdf)")
        if not file_path:
            return

        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Barokah Copy & Printing", styles["Title"]),
            Paragraph("Laporan Sejarah Transaksi", styles["Heading2"]),
            Paragraph(
                f"Periode: {self.current_filters.get('date_from', '-')} s/d {self.current_filters.get('date_to', '-')}",
                styles["BodyText"],
            ),
            Spacer(1, 10),
        ]

        stats = self.current_statistics or {}
        top_cashier = stats.get("top_cashier", {})
        summary_text = (
            f"Total Transaksi: {stats.get('total_count', 0)} | "
            f"Total Revenue: {self._format_currency(stats.get('total_revenue', 0))} | "
            f"Rata-rata: {self._format_currency(stats.get('avg_transaction', 0))} | "
            f"Kasir Teratas: {top_cashier.get('name', '-')} ({top_cashier.get('count', 0)} trx)"
        )
        story.append(Paragraph(summary_text, styles["BodyText"]))
        story.append(Spacer(1, 12))

        table_data = [["ID", "Tanggal", "Kasir", "Customer", "Items", "Total", "Metode"]]
        for transaction in transactions_data:
            table_data.append([
                str(transaction.get("id", "-")),
                self._format_datetime(transaction.get("tanggal")),
                transaction.get("kasir_name", "-"),
                transaction.get("customer_name", "-"),
                transaction.get("items_preview", "-"),
                self._format_currency(transaction.get("total", 0)),
                transaction.get("payment_method", "-"),
            ])

        table = Table(table_data, repeatRows=1, colWidths=[45, 100, 90, 100, 250, 90, 80])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#AAB7C4")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FBFF")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 10),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M:%S')}", styles["BodyText"]))
        doc.build(story)
        QMessageBox.information(self, "Export PDF", f"File PDF berhasil disimpan ke:\n{file_path}")

    def print_transaction(self, transaction_id):
        detail = self.fetch_transaction_detail(transaction_id)
        if not detail:
            QMessageBox.warning(self, "Print", "Detail transaksi tidak ditemukan.")
            return

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setPageMargins(6, 6, 6, 6, QPrinter.Unit.Millimeter)
        printer.setDocName(f"Transaksi_{transaction_id}")

        dialog = QPrintDialog(printer, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        header = detail.get("header", {})
        lines = [
            "Barokah Copy & Printing",
            f"Transaksi #{header.get('id', '-')}",
            self._format_datetime(header.get("tanggal")),
            f"Kasir   : {header.get('kasir_name', '-')}",
            f"Customer: {header.get('customer_name', '-')}",
            f"Metode  : {header.get('payment_method', '-')}",
            "-" * 32,
        ]
        for item in detail.get("items", []):
            lines.append(f"{item.get('product_name', '-')}")
            lines.append(
                f"{item.get('quantity', 0)} x {self._format_currency(item.get('price', 0))} = {self._format_currency(item.get('subtotal', 0))}"
            )
        lines.extend([
            "-" * 32,
            f"Subtotal : {self._format_currency(header.get('subtotal', 0))}",
            f"Diskon   : {self._format_currency(header.get('diskon_nominal', 0))}",
            f"Total    : {self._format_currency(header.get('total', 0))}",
            f"Bayar    : {self._format_currency(header.get('amount_paid', 0))}",
            f"Kembali  : {self._format_currency(header.get('change_amount', 0))}",
            "Terima kasih atas kunjungan Anda",
        ])

        document = QTextDocument()
        document.setDefaultFont(QFont("Courier New", 10))
        document.setPlainText("\n".join(lines))
        document.print(printer)

    def _handle_print_action(self):
        selected_items = self.transaction_table.selectedItems()
        if selected_items:
            transaction_id = selected_items[0].data(Qt.ItemDataRole.UserRole)
            if transaction_id:
                self.print_transaction(transaction_id)
                return

        self._print_transactions_summary(self.current_transactions)

    def _print_transactions_summary(self, transactions_data):
        if not transactions_data:
            QMessageBox.information(self, "Print", "Tidak ada transaksi untuk dicetak.")
            return

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOrientation(QPrinter.Orientation.Landscape)
        dialog = QPrintDialog(printer, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        summary_lines = [
            "LAPORAN SEJARAH TRANSAKSI",
            f"Periode: {self.current_filters.get('date_from', '-')} s/d {self.current_filters.get('date_to', '-')}",
            "-" * 70,
        ]
        for transaction in transactions_data:
            summary_lines.append(
                f"#{transaction.get('id')} | {self._format_datetime(transaction.get('tanggal'))} | "
                f"{transaction.get('kasir_name', '-')} | {transaction.get('customer_name', '-')} | "
                f"{self._format_currency(transaction.get('total', 0))} | {transaction.get('payment_method', '-')}"
            )

        document = QTextDocument()
        document.setDefaultFont(QFont("Courier New", 10))
        document.setPlainText("\n".join(summary_lines))
        document.print(printer)

    def _build_export_filename(self, extension):
        start_date = self.current_filters.get("date_from", self.date_from_input.date().toString("yyyy-MM-dd"))
        end_date = self.current_filters.get("date_to", self.date_to_input.date().toString("yyyy-MM-dd"))
        return str(Path.cwd() / f"transaksi_{start_date}_{end_date}.{extension}")

    @staticmethod
    def _format_currency(value):
        amount = int(float(value or 0))
        return f"Rp {amount:,}".replace(",", ".")

    @staticmethod
    def _format_datetime(value):
        if not value:
            return "-"
        text = str(value)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
            try:
                parsed = datetime.strptime(text, fmt)
                return parsed.strftime("%d %b %Y, %H:%M")
            except ValueError:
                continue
        return text
