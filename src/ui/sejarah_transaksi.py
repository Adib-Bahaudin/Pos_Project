import sqlite3
from datetime import datetime
from config import asset_path

import openpyxl
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QColor, QPixmap, QTextCharFormat
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QDateEdit, QLineEdit, QFrame, QFileDialog
)
from openpyxl.styles import Font

HAS_OPENPYXL = True

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
HAS_REPORTLAB = True

from src.database.database import DatabaseManager
from src.utils.fungsi import CustomCalendar, NavigationButton, CurrencyDelegate
from src.ui.transaction_detail import TransactionDetailModal
from src.utils.message import CustomMessageBox


class SejarahTransaksiWindow(QWidget):
    def __init__(self, user_data=None):
        super().__init__()
        self.user_data = user_data or {}
        self.user_role = self.user_data.get("role", "")
        self.db = DatabaseManager()
        
        self.current_page = 1
        self.items_per_page = 9
        self.transactions_data = []
        
        self._setup_ui()
        self._load_cashiers()
        self.apply_filters()
        
    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        
        header_label = QLabel("Sejarah Transaksi")
        header_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        header_label.setStyleSheet("color: white; background-color: transparent")
        main_layout.addWidget(header_label)
        
        self.stats_panel = self._create_stats_panel()
        main_layout.addWidget(self.stats_panel)
        
        self.filter_panel = self._create_filter_panel()
        main_layout.addWidget(self.filter_panel)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.setStyleSheet("""
            QTableWidget { 
                background-color: #2d2d2d; 
                alternate-background-color: #545454; 
                color: white; 
                border: none; 
                outline: none; 
            }
            QTableWidget::item:selected {
                background-color: #0078D7;
                color: white;
            }
            QHeaderView::section:horizontal {
                background-color: #a6a6a6;
                color: black; 
                font-weight: bold;
                border: none; 
                padding: 6px; 
            }
        """)

        self.table.setColumnCount(10) 
        self.table.setHorizontalHeaderLabels([
            "No", "ID", "Tanggal", "Kasir", "Customer", 
            "Metode", "Subtotal", "Diskon", "Pembulatan", "Total"
        ])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.cellDoubleClicked.connect(self.show_transaction_detail_modal)

        currency_delegate = CurrencyDelegate(horizontal_padding=25, parent=self.table)
        self.table.setItemDelegateForColumn(6, currency_delegate)
        self.table.setItemDelegateForColumn(7, currency_delegate)
        self.table.setItemDelegateForColumn(8, currency_delegate)
        self.table.setItemDelegateForColumn(9, currency_delegate)

        self.table.setRowCount(self.items_per_page)
        row_height = 30
        self.table.verticalHeader().setDefaultSectionSize(row_height)
        header_height = self.table.horizontalHeader().height()
        self.table.setFixedHeight(header_height + self.items_per_page * row_height + 2)

        main_layout.addWidget(self.table)
        
        pagination_layout = QHBoxLayout()
        self.btn_prev = NavigationButton(asset_path("arah kiri.svg"), asset_path("kiri-hover.svg"))
        self.btn_next = NavigationButton(asset_path("arah kanan.svg"), asset_path("kanan-hover.svg"))

        self.lbl_page = QLabel(f"Halaman {self.current_page}")
        self.lbl_page.setStyleSheet("color: white; background-color: transparent;")
        
        self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next.clicked.connect(self._next_page)

        pagination_layout.addWidget(self.btn_prev)
        pagination_layout.addWidget(self.lbl_page)
        pagination_layout.addWidget(self.btn_next)
        pagination_layout.addStretch()
        
        main_layout.addLayout(pagination_layout)
        
        action_layout = QHBoxLayout()
        self.btn_export_excel = QPushButton("📥 Export Excel")
        self.btn_export_pdf = QPushButton("📄 Export PDF")
        
        btn_style = "background-color: #0078D7; color: white; padding: 8px; border-radius: 4px;"
        self.btn_export_excel.setStyleSheet(btn_style)
        self.btn_export_pdf.setStyleSheet(btn_style)
        
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        self.btn_export_pdf.clicked.connect(self.export_to_pdf)
        
        action_layout.addStretch()
        action_layout.addWidget(self.btn_export_excel)
        action_layout.addWidget(self.btn_export_pdf)
        
        main_layout.addLayout(action_layout)
        
        self.setStyleSheet("background-color: #1e1e1e;")

    def _create_stats_panel(self):
        frame = QFrame()
        frame.setStyleSheet("background-color: transparent; border-radius: 8px;")
        layout = QHBoxLayout(frame)
        
        self.lbl_stat_count = self._create_stat_card("Total Transaksi", "0", "#b174e7", asset_path("ikon_keranjang.svg"))
        self.lbl_stat_revenue = self._create_stat_card("Total Revenue", "Rp 0", "#5271ff", asset_path("ikon_koin.svg"))
        self.lbl_stat_avg = self._create_stat_card("Rata-rata", "Rp 0", "#ff914d", asset_path("ikon_rata-rata.svg"))
        self.lbl_stat_top = self._create_stat_card("Kasir Teratas", "-", "#ffde59", asset_path("ikon_medali.svg"))
        
        layout.addWidget(self.lbl_stat_count)
        layout.addWidget(self.lbl_stat_revenue)
        layout.addWidget(self.lbl_stat_avg)
        layout.addWidget(self.lbl_stat_top)
        
        return frame

    @staticmethod
    def _create_stat_card(title, initial_value, base_color, image_path):
        container = QFrame()
        container.setObjectName("statCard") 
        
        color = QColor(base_color)
        border_color = color.darker(200).name()
        
        style = f"""
            QFrame#statCard {{
                background: qlineargradient(x1: 1, y1: 0, x2: 0, y2: 1, 
                                            stop: 0 {base_color}, stop: 1 black);
                border: 2px solid {border_color};
                border-radius: 4px;
            }}
        """
        container.setStyleSheet(style)
        
        layout = QVBoxLayout(container)
        layout.setContentsMargins(10, 10, 10, 10)
        
        title_layout = QHBoxLayout()
        
        lbl_title = QLabel(title)
        lbl_title.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        lbl_title.setStyleSheet("color: white; font-size: 15px; border: none; background: transparent;")
        
        lbl_image = QLabel()
        pixmap = QPixmap(image_path)
        pixmap = pixmap.scaled(24, 24, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        lbl_image.setPixmap(pixmap)
        lbl_image.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        lbl_image.setStyleSheet("border: none; background: transparent;")
        
        title_layout.addWidget(lbl_title)
        title_layout.addWidget(lbl_image)
        
        lbl_value = QLabel(initial_value)
        lbl_value.setObjectName("statValue")
        lbl_value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_value.setStyleSheet("color: white; font-size: 18px; font-weight: bold; border: none; background: transparent;")
        
        layout.addLayout(title_layout)
        layout.addWidget(lbl_value)
        
        return container

    @staticmethod
    def _update_stat_card(card_container, new_value):
        lbl_value = card_container.findChild(QLabel, "statValue")
        if lbl_value:
            lbl_value.setText(new_value)
        else:
            labels = card_container.findChildren(QLabel)
            if len(labels) >= 3:
                labels[2].setText(new_value)

    def _create_filter_panel(self):
        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #545454; border-radius: 8px; }")
        layout = QHBoxLayout(frame)
        
        # --- STYLESHEET KHUSUS DATE EDIT & CALENDAR POPUP ---
        date_style = """
            QDateEdit {
                background-color: #1e1e1e; 
                color: white;
                border: 1px solid #444;
                border-radius: 4px;
                padding: 4px;
            }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 24px;
                border-left: 1px solid #444;
            }
            /* --- FIX: Menggunakan Ikon Custom --- */
            QDateEdit::down-arrow {
                image: url(data/icon_down.svg);
                width: 12px;  /* Sesuaikan ukuran ikon jika dirasa kurang besar/kecil */
                height: 12px;
            }
            QCalendarWidget QWidget {
                background-color: #1e1e1e;
                color: white;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #2b2b2b;
                border-bottom: 1px solid #444;
                min-height: 30px;
            }
            QCalendarWidget QToolButton {
                color: white;
                background-color: transparent;
                font-weight: bold;
                padding: 4px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover {
                background-color: #555;
                border-radius: 4px;
            }
            QCalendarWidget QToolButton::menu-indicator {
                image: none;
            }
            QMenu {
                background-color: #2b2b2b;
                color: white;
                border: 1px solid #555;
            }
            QMenu::item {
                background-color: transparent;
                color: white;
                padding: 6px 20px;
            }
            QMenu::item:selected {
                background-color: #0078D7;
                color: white;
            }
            QCalendarWidget QTableView {
                background-color: #1e1e1e;
                color: white;
                selection-background-color: #0078D7;
                selection-color: white;
            }
            QCalendarWidget QSpinBox {
                background-color: #1e1e1e;
                color: white;
                border: 1px solid #444;
            }
        """

        # --- STYLESHEET KHUSUS COMBOBOX ---
        combo_style = """
            QComboBox {
                background-color: #1e1e1e; 
                color: white; 
                padding: 4px 8px; /* Padding sedikit lebih lebar agar teks tidak menempel border */
                border: 1px solid #444;
                border-radius: 4px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 24px;
                border-left: 1px solid #444;
            }
            /* --- FIX: Menggunakan Ikon Custom --- */
            QComboBox::down-arrow {
                image: url(data/icon_down.svg);
                width: 12px;
                height: 12px;
            }
            /* Styling list dropdown saat Combo Box diklik agar senada */
            QComboBox QAbstractItemView {
                background-color: #2b2b2b;
                color: white;
                border: 1px solid #444;
                selection-background-color: #0078D7;
            }
        """

        header_format = QTextCharFormat()
        header_format.setBackground(QColor("#2b2b2b"))
        header_format.setForeground(QColor("#A9B7C6"))
        
        self.date_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        self.date_from.setStyleSheet(date_style)

        cal_from = CustomCalendar()
        cal_from.setHeaderTextFormat(header_format)
        self.date_from.setCalendarWidget(cal_from)
        
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setStyleSheet(date_style)

        cal_to = CustomCalendar()
        cal_to.setHeaderTextFormat(header_format)
        self.date_to.setCalendarWidget(cal_to)
        
        self.cb_kasir = QComboBox()
        # Terapkan stylesheet combo_style
        self.cb_kasir.setStyleSheet(combo_style)
        
        self.cb_metode = QComboBox()
        self.cb_metode.addItems(["Semua", "Tunai", "Transfer", "QRIS"])
        # Terapkan stylesheet combo_style
        self.cb_metode.setStyleSheet(combo_style)
        
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Cari ID/Customer...")
        self.search_box.setStyleSheet("background-color: #1e1e1e; color: white; padding: 4px; border: 1px solid #444; border-radius: 4px;")
        
        self.btn_filter = QPushButton("Filter")
        self.btn_filter.setStyleSheet("background-color: #28A745; color: white; padding: 6px 15px; border-radius: 4px; font-weight: bold;")
        self.btn_filter.clicked.connect(lambda: self.apply_filters(reset_page=True))
        
        lbl_dari = QLabel("Dari:")
        lbl_dari.setStyleSheet("color: white;")
        layout.addWidget(lbl_dari)
        layout.addWidget(self.date_from)
        lbl_sampai = QLabel("Sampai:")
        lbl_sampai.setStyleSheet("color: white;")
        layout.addWidget(lbl_sampai)
        layout.addWidget(self.date_to)
        lbl_kasir = QLabel("Kasir:")
        lbl_kasir.setStyleSheet("color: white;")
        layout.addWidget(lbl_kasir)
        layout.addWidget(self.cb_kasir)
        lbl_metode = QLabel("Metode:")
        lbl_metode.setStyleSheet("color: white;")
        layout.addWidget(lbl_metode)
        layout.addWidget(self.cb_metode)
        layout.addWidget(self.search_box)
        layout.addWidget(self.btn_filter)
        
        return frame

    def _load_cashiers(self):
        conn = sqlite3.connect(self.db.db_name)
        cursor = conn.cursor()
        cursor.execute("SELECT id, nama FROM users")
        kasir_list = cursor.fetchall()
        conn.close()
        
        self.cb_kasir.addItem("Semua", None)
        for k in kasir_list:
            self.cb_kasir.addItem(k[1], k[0])
            
        if self.user_role != "Super_user":
            idx = self.cb_kasir.findText(self.user_data.get("username", ""))
            if idx >= 0:
                self.cb_kasir.setCurrentIndex(idx)
            self.cb_kasir.setEnabled(False)

    def _get_filter_values(self):
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        kasir_id = self.cb_kasir.currentData()
        payment_method = self.cb_metode.currentText()
        search_keyword = self.search_box.text()
        
        return {
            "date_from": date_from,
            "date_to": date_to,
            "kasir_id": kasir_id,
            "payment_method": payment_method,
            "search_keyword": search_keyword,
            "amount_min": None,
            "amount_max": None
        }

    def apply_filters(self, reset_page=False):
        if reset_page:
            self.current_page = 1
            
        filters = self._get_filter_values()
        offset = (self.current_page - 1) * self.items_per_page
        
        self.transactions_data = self.db.get_transaction_history(filters, self.items_per_page, offset)
        self.populate_table()
        
        stats = self.db.get_transaction_statistics(filters)
        self._update_stat_card(self.lbl_stat_count, str(stats.get('total_count', 0)))
        
        rev = stats.get('total_revenue', 0)
        self._update_stat_card(self.lbl_stat_revenue, f"Rp {int(rev):,}")
        
        avg = stats.get('avg_transaction', 0)
        self._update_stat_card(self.lbl_stat_avg, f"Rp {int(avg):,}")
        
        top = stats.get('top_cashier', '-')
        top_c = stats.get('top_cashier_count', 0)
        self._update_stat_card(self.lbl_stat_top, f"{top} ({top_c})")
        
        self.lbl_page.setText(f"Halaman {self.current_page}")
        self.btn_prev.setEnabled(self.current_page > 1)
        self.btn_next.setEnabled(len(self.transactions_data) == self.items_per_page)

    def populate_table(self):
        self.table.clearContents()

        def create_center_item(text):
            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            return item

        for i, row in enumerate(self.transactions_data):
            
            dt_str = row.get("tanggal", "")
            try:
                if len(dt_str) > 10:
                    dt_obj = datetime.fromisoformat(dt_str)
                    dt_disp = dt_obj.strftime("%d %b %Y, %H:%M")
                else:
                    dt_disp = dt_str
            except Exception:
                dt_disp = dt_str

            self.table.setItem(i, 0, create_center_item(str(i + 1 + (self.current_page - 1) * self.items_per_page)))
            self.table.setItem(i, 1, create_center_item(str(row.get("id", ""))))
            self.table.setItem(i, 2, create_center_item(dt_disp))
            self.table.setItem(i, 3, create_center_item(row.get("nama_kasir", "")))
            self.table.setItem(i, 4, create_center_item(row.get("nama_customer", "")))
            
            metode = row.get("metode_bayar", "")
            item_metode = create_center_item(metode)
            if "tunai" in metode.lower():
                item_metode.setForeground(QColor("#28A745"))
            elif "transfer" in metode.lower():
                item_metode.setForeground(QColor("#0078D7"))
            else:
                item_metode.setForeground(QColor("#FFC107"))
                
            self.table.setItem(i, 5, item_metode)

            self.table.setItem(i, 6, QTableWidgetItem(f"Rp. {int(row.get('subtotal', 0)):,}"))
            self.table.setItem(i, 7, QTableWidgetItem(f"Rp. {int(row.get('diskon_nominal', 0)):,}"))

            pembulatan = int(row.get('pembulatan', 0))
            item_pembulatan = QTableWidgetItem(self._format_pembulatan(pembulatan))
            if pembulatan > 0:
                item_pembulatan.setForeground(QColor("#28A745"))
            elif pembulatan < 0:
                item_pembulatan.setForeground(QColor("#FF6B6B"))
            self.table.setItem(i, 8, item_pembulatan)

            self.table.setItem(i, 9, QTableWidgetItem(f"Rp. {int(row.get('total', 0)):,}"))

    @staticmethod
    def _format_pembulatan(nilai: int) -> str:
        if nilai > 0:
            return f"+Rp. {nilai:,}"
        elif nilai < 0:
            return f"-Rp. {abs(nilai):,}"
        return "Rp. 0"

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.apply_filters()

    def _next_page(self):
        if len(self.transactions_data) == self.items_per_page:
            self.current_page += 1
            self.apply_filters()

    def show_transaction_detail_modal(self, row):
        if row < 0 or row >= len(self.transactions_data):
            return
        t_id = self.transactions_data[row].get("id")
        if t_id:
            dialog = TransactionDetailModal(self.db, t_id, self)
            dialog.exec()

    def export_to_excel(self):
        if not HAS_OPENPYXL:
            CustomMessageBox.warning(self, "Error", "openpyxl belum terinstall. Silakan 'pip install -r requirements.txt'")
            return
            
        filters = self._get_filter_values()
        all_data = self.db.get_transaction_history(filters, limit=10000)
        if not all_data:
            CustomMessageBox.information(self, "Info", "Tidak ada data untuk diexport.")
            return
            
        default_name = f"transaksi_{filters['date_from']}_{filters['date_to']}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", default_name, "Excel Files (*.xlsx)")
        if not path:
            return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Sejarah Transaksi"
            
                headers = ["ID", "Tanggal", "ID Kasir", "Nama Kasir", "Customer", "Subtotal", "Diskon", "Pembulatan", "Total", "Metode Bayar", "Catatan"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = Font(bold=True)
                
                for row_idx, data in enumerate(all_data, 2):
                    ws.cell(row=row_idx, column=1, value=data.get("id"))
                    ws.cell(row=row_idx, column=2, value=data.get("tanggal"))
                    ws.cell(row=row_idx, column=3, value=data.get("id_kasir"))
                    ws.cell(row=row_idx, column=4, value=data.get("nama_kasir"))
                    ws.cell(row=row_idx, column=5, value=data.get("nama_customer"))
                    ws.cell(row=row_idx, column=6, value=int(data.get("subtotal", 0)))
                    ws.cell(row=row_idx, column=7, value=int(data.get("diskon_nominal", 0)))
                    ws.cell(row=row_idx, column=8, value=int(data.get("pembulatan", 0)))
                    ws.cell(row=row_idx, column=9, value=int(data.get("total", 0)))
                    ws.cell(row=row_idx, column=10, value=data.get("metode_bayar"))
                    ws.cell(row=row_idx, column=11, value=data.get("catatan"))
            else:
                CustomMessageBox.critical(self, "Error", "Gagal membuat worksheet Excel")
                
            wb.save(path)
            CustomMessageBox.information(self, "Sukses", f"Data berhasil diexport ke:\n{path}")
        except Exception as e:
            CustomMessageBox.critical(self, "Error", f"Gagal export Excel: {str(e)}")

    def export_to_pdf(self):
        if not HAS_REPORTLAB:
            CustomMessageBox.warning(self, "Error", "reportlab belum terinstall. Silakan 'pip install -r requirements.txt'")
            return
            
        filters = self._get_filter_values()
        all_data = self.db.get_transaction_history(filters, limit=10000)
        if not all_data:
            CustomMessageBox.information(self, "Info", "Tidak ada data untuk diexport.")
            return
            
        default_name = f"transaksi_{filters['date_from']}_{filters['date_to']}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", default_name, "PDF Files (*.pdf)")
        if not path:
            return
            
        try:
            doc = SimpleDocTemplate(path, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            title = Paragraph(f"Laporan Transaksi Kasir ({filters['date_from']} s/d {filters['date_to']})", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            table_data = [["ID", "Tanggal", "Kasir", "Customer", "Subtotal", "Diskon", "Pembulatan", "Total", "Metode"]]
            for d in all_data:
                pembulatan_val = int(d.get('pembulatan', 0))
                pembulatan_str = f"+Rp {pembulatan_val:,}" if pembulatan_val > 0 else (f"-Rp {abs(pembulatan_val):,}" if pembulatan_val < 0 else "Rp 0")
                table_data.append([
                    str(d.get("id", "")),
                    str(d.get("tanggal", "")),
                    str(d.get("nama_kasir", "")),
                    str(d.get("nama_customer", "")),
                    f"Rp {int(d.get('subtotal', 0)):,}",
                    f"Rp {int(d.get('diskon_nominal', 0)):,}",
                    pembulatan_str,
                    f"Rp {int(d.get('total', 0)):,}",
                    str(d.get("metode_bayar", ""))
                ])
                
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            elements.append(t)
            doc.build(elements)
            
            CustomMessageBox.information(self, "Sukses", f"Data berhasil diexport ke:\n{path}")
        except Exception as e:
            CustomMessageBox.critical(self, "Error", f"Gagal export PDF: {str(e)}")

    def refresh_data(self):
        """
        Mengembalikan semua filter ke kondisi awal (default) 
        dan mengambil data transaksi terbaru dari database.
        """
        self.current_page = 1
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_to.setDate(QDate.currentDate())
        self.search_box.clear()
        self.cb_metode.setCurrentIndex(0)
        self.cb_kasir.clear()
        self._load_cashiers()
        self.apply_filters(reset_page=True)
